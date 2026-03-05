import argparse
import json
import re
from pathlib import Path

import openpyxl

MONTH_COLS = list(range(4, 16))  # D:O
ORG_MAP = {
    "西部": "west",
    "多摩": "tama",
    "神奈川": "kanagawa",
    "湘南": "shonan",
}
ITEM_MAP = {
    "総売上目標": "totalTarget",
    "杭目標": "pileTarget",
    "設計目標": "designTarget",
    "総売上": "totalActual",
    "杭売上": "pileActual",
    "設計売上": "designActual",
}


def to_num(v):
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except Exception:
        return 0


def extract_rows(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    result = {k: {} for k in ORG_MAP.values()}
    current_org = None

    for r in range(1, ws.max_row + 1):
        org_cell = ws.cell(r, 2).value
        item_cell = ws.cell(r, 3).value

        if isinstance(org_cell, str):
            if org_cell in ORG_MAP:
                current_org = ORG_MAP[org_cell]
            elif org_cell.strip():
                current_org = None

        if not current_org or not isinstance(item_cell, str):
            continue

        if item_cell not in ITEM_MAP:
            continue

        key = ITEM_MAP[item_cell]
        values = [to_num(ws.cell(r, c).value) for c in MONTH_COLS]
        result[current_org][key] = values

    missing = []
    for org_key, row in result.items():
        for needed in ITEM_MAP.values():
            if needed not in row:
                missing.append(f"{org_key}:{needed}")

    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(f"必要なデータが不足しています: {joined}")

    return result


def load_existing_terms(out_path: Path):
    if not out_path.exists():
        return {}

    text = out_path.read_text(encoding="utf-8")
    m = re.search(r"window\.EXCEL_TERMS\s*=\s*(\{[\s\S]*?\})\s*;", text)
    if not m:
        return {}

    try:
        return json.loads(m.group(1))
    except Exception:
        return {}


def main():
    parser = argparse.ArgumentParser(description="目標比較シートから data.js を生成（複数期保持）")
    parser.add_argument("--xlsx", required=True, help="入力Excelファイル")
    parser.add_argument("--sheet", default="目標比較", help="対象シート名")
    parser.add_argument("--out", default="data.js", help="出力JSファイル")
    parser.add_argument("--term-no", type=int, required=True, help="期番号 (例: 39)")
    parser.add_argument("--start-year", type=int, required=True, help="期開始年 (例: 2025)")
    args = parser.parse_args()

    out_path = Path(args.out)
    terms = load_existing_terms(out_path)

    terms[str(args.term_no)] = {
        "termNo": args.term_no,
        "startYear": args.start_year,
        "rows": extract_rows(Path(args.xlsx), args.sheet),
    }

    latest_key = sorted(terms.keys(), key=lambda k: int(k))[-1]
    latest = terms[latest_key]

    out = (
        "window.EXCEL_TERMS = "
        + json.dumps(terms, ensure_ascii=False, indent=2)
        + ";\n\n"
        + "window.EXCEL_FISCAL_META = "
        + json.dumps({"termNo": latest["termNo"], "startYear": latest["startYear"]}, ensure_ascii=False)
        + ";\n\n"
        + "window.EXCEL_TARGET_COMPARISON = "
        + json.dumps(latest["rows"], ensure_ascii=False, indent=2)
        + ";\n"
    )

    out_path.write_text(out, encoding="utf-8")
    print(f"generated: {args.out} (terms: {', '.join(sorted(terms.keys(), key=lambda x: int(x)))})")


if __name__ == "__main__":
    main()
